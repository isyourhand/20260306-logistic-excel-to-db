from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


EXCEL_PATTERNS = ("*.xlsx", "*.xlsm", "*.xltx", "*.xltm")
INVALID_FS_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def configure_console_encoding() -> None:
    # Avoid UnicodeEncodeError on Windows GBK consoles when filenames include
    # characters such as NBSP (U+00A0).
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def safe_name(name: str) -> str:
    """Make a filesystem-safe path segment while preserving readability."""
    cleaned = INVALID_FS_CHARS.sub("_", name).strip()
    return cleaned or "unnamed"


def iter_excel_files(input_dir: Path) -> Iterable[Path]:
    for pattern in EXCEL_PATTERNS:
        for path in input_dir.glob(pattern):
            if path.is_file() and not path.name.startswith("~$"):
                yield path


def cell_to_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def is_meaningful(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def infer_effective_bounds(worksheet) -> tuple[int, int]:
    """
    Infer a practical grid size from meaningful cells.
    This avoids sparse-matrix explosions caused by accidental formatting far away.
    """
    max_row = 1
    max_col = 1
    has_meaningful = False

    for (row, col), cell in worksheet._cells.items():
        if is_meaningful(cell.value):
            has_meaningful = True
            if row > max_row:
                max_row = row
            if col > max_col:
                max_col = col

    for merged in worksheet.merged_cells.ranges:
        top_left = worksheet.cell(merged.min_row, merged.min_col).value
        if is_meaningful(top_left):
            has_meaningful = True
            if merged.max_row > max_row:
                max_row = merged.max_row
            if merged.max_col > max_col:
                max_col = merged.max_col

    if not has_meaningful:
        return 1, 1
    return max_row, max_col


def build_fill_ranges(worksheet, max_row: int):
    row_ranges = defaultdict(list)
    for merged in worksheet.merged_cells.ranges:
        value = cell_to_text(worksheet.cell(merged.min_row, merged.min_col).value)
        start_row = merged.min_row
        end_row = min(merged.max_row, max_row)
        for row in range(start_row, end_row + 1):
            row_ranges[row].append((merged.min_col, merged.max_col, value))
    return row_ranges


def fill_value_from_ranges(col: int, ranges, raw_value: str) -> str:
    for min_col, max_col, merged_value in ranges:
        if min_col <= col <= max_col:
            return merged_value
    return raw_value


def write_merged_ranges_json(path: Path, worksheet):
    payload = {
        "sheet": worksheet.title,
        "merged_ranges": [str(rng) for rng in worksheet.merged_cells.ranges],
    }
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def export_sheet_csvs(worksheet, ws_dir: Path, max_row: int, max_col: int):
    fill_ranges = build_fill_ranges(worksheet, max_row)

    with (ws_dir / "grid.csv").open("w", newline="", encoding="utf-8-sig") as f_grid, (
        ws_dir / "grid_filled.csv"
    ).open("w", newline="", encoding="utf-8-sig") as f_filled:
        grid_writer = csv.writer(f_grid)
        filled_writer = csv.writer(f_filled)

        for r in range(1, max_row + 1):
            row_raw = []
            row_filled = []
            row_ranges = fill_ranges.get(r, [])

            for c in range(1, max_col + 1):
                raw_value = cell_to_text(worksheet.cell(row=r, column=c).value)
                row_raw.append(raw_value)
                row_filled.append(fill_value_from_ranges(c, row_ranges, raw_value))

            grid_writer.writerow(row_raw)
            filled_writer.writerow(row_filled)


def process_workbook(path: Path, output_root: Path, data_only: bool = False, bounds_mode: str = "effective"):
    workbook = load_workbook(path, data_only=data_only)
    processed = []

    for sheet_idx, ws in enumerate(workbook.worksheets, start=1):
        print(f"    - [{sheet_idx}/{len(workbook.worksheets)}] {ws.title}", flush=True)
        wb_dir = output_root / safe_name(path.stem)
        ws_dir = wb_dir / safe_name(ws.title)
        ws_dir.mkdir(parents=True, exist_ok=True)

        if bounds_mode == "strict":
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
        else:
            max_row, max_col = infer_effective_bounds(ws)

        print(f"      bounds: rows={max_row}, cols={max_col}", flush=True)

        write_merged_ranges_json(ws_dir / "merged_ranges.json", ws)
        export_sheet_csvs(ws, ws_dir, max_row, max_col)

        processed.append(ws.title)

    workbook.close()
    return processed


def main():
    configure_console_encoding()
    parser = argparse.ArgumentParser(
        description=(
            "Export each sheet in every Excel workbook to structured grids: "
            "grid.csv, merged_ranges.json, and grid_filled.csv"
        )
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path.cwd(),
        help="Directory containing Excel files (default: current directory)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path.cwd() / "out",
        help="Output directory root (default: ./out)",
    )
    parser.add_argument(
        "--data-only",
        action="store_true",
        help="Read calculated values for formula cells instead of formula text",
    )
    parser.add_argument(
        "--bounds-mode",
        choices=("effective", "strict"),
        default="effective",
        help=(
            "effective: infer practical bounds from meaningful cells (default); "
            "strict: use worksheet max_row/max_column exactly"
        ),
    )

    args = parser.parse_args()
    input_dir: Path = args.input_dir.resolve()
    output_dir: Path = args.output_dir.resolve()

    excel_files = sorted(set(iter_excel_files(input_dir)), key=lambda p: p.name.lower())
    if not excel_files:
        print(f"No Excel files found in: {input_dir}")
        return

    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Input directory : {input_dir}")
    print(f"Output directory: {output_dir}")
    print(f"Found {len(excel_files)} workbook(s).")

    for idx, excel_file in enumerate(excel_files, start=1):
        print(f"[{idx}/{len(excel_files)}] Processing: {excel_file.name}")
        try:
            sheets = process_workbook(
                excel_file,
                output_dir,
                data_only=args.data_only,
                bounds_mode=args.bounds_mode,
            )
            print(f"    Exported sheets: {', '.join(sheets)}")
        except Exception as exc:
            print(f"    Skipped due to error: {exc}")

    print("Done.")


if __name__ == "__main__":
    main()
