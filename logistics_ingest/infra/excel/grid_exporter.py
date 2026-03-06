from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from logistics_ingest.shared.settings import EXCEL_PATTERNS
from logistics_ingest.shared.text_utils import safe_name

__all__ = [
    "EXCEL_PATTERNS",
    "iter_excel_files",
    "process_workbook",
    "safe_name",
]


def iter_excel_files(input_dir: Path) -> Iterable[Path]:
    for pattern in EXCEL_PATTERNS:
        for path in input_dir.glob(pattern):
            if path.is_file() and not path.name.startswith("~$"):
                yield path


def cell_to_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def is_meaningful(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def infer_effective_bounds(worksheet) -> tuple[int, int]:
    max_row = 1
    max_col = 1
    has_meaningful = False

    for (row, col), cell in worksheet._cells.items():
        if is_meaningful(cell.value):
            has_meaningful = True
            if row > max_row:
                max_row = row
            if col > max_col:
                max_col = col

    for merged in worksheet.merged_cells.ranges:
        top_left = worksheet.cell(merged.min_row, merged.min_col).value
        if is_meaningful(top_left):
            has_meaningful = True
            if merged.max_row > max_row:
                max_row = merged.max_row
            if merged.max_col > max_col:
                max_col = merged.max_col

    if not has_meaningful:
        return 1, 1
    return max_row, max_col


def build_fill_ranges(worksheet, max_row: int):
    row_ranges = defaultdict(list)
    for merged in worksheet.merged_cells.ranges:
        value = cell_to_text(worksheet.cell(merged.min_row, merged.min_col).value)
        start_row = merged.min_row
        end_row = min(merged.max_row, max_row)
        for row in range(start_row, end_row + 1):
            row_ranges[row].append((merged.min_col, merged.max_col, value))
    return row_ranges


def fill_value_from_ranges(col: int, ranges, raw_value: str) -> str:
    for min_col, max_col, merged_value in ranges:
        if min_col <= col <= max_col:
            return merged_value
    return raw_value


def write_merged_ranges_json(path: Path, worksheet):
    payload = {
        "sheet": worksheet.title,
        "merged_ranges": [str(rng) for rng in worksheet.merged_cells.ranges],
    }
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def export_sheet_csvs(worksheet, ws_dir: Path, max_row: int, max_col: int):
    fill_ranges = build_fill_ranges(worksheet, max_row)

    with (ws_dir / "grid.csv").open("w", newline="", encoding="utf-8-sig") as f_grid, (
        ws_dir / "grid_filled.csv"
    ).open("w", newline="", encoding="utf-8-sig") as f_filled:
        grid_writer = csv.writer(f_grid)
        filled_writer = csv.writer(f_filled)

        for r in range(1, max_row + 1):
            row_raw = []
            row_filled = []
            row_ranges = fill_ranges.get(r, [])

            for c in range(1, max_col + 1):
                raw_value = cell_to_text(worksheet.cell(row=r, column=c).value)
                row_raw.append(raw_value)
                row_filled.append(fill_value_from_ranges(c, row_ranges, raw_value))

            grid_writer.writerow(row_raw)
            filled_writer.writerow(row_filled)


def process_workbook(path: Path, output_root: Path, data_only: bool = False, bounds_mode: str = "effective"):
    workbook = load_workbook(path, data_only=data_only)
    processed = []

    for sheet_idx, ws in enumerate(workbook.worksheets, start=1):
        print(f"    - [{sheet_idx}/{len(workbook.worksheets)}] {ws.title}", flush=True)
        wb_dir = output_root / safe_name(path.stem)
        ws_dir = wb_dir / safe_name(ws.title)
        ws_dir.mkdir(parents=True, exist_ok=True)

        if bounds_mode == "strict":
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
        else:
            max_row, max_col = infer_effective_bounds(ws)

        print(f"      bounds: rows={max_row}, cols={max_col}", flush=True)

        write_merged_ranges_json(ws_dir / "merged_ranges.json", ws)
        export_sheet_csvs(ws, ws_dir, max_row, max_col)

        processed.append(ws.title)

    workbook.close()
    return processed
