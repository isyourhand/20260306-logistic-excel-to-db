from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def read_workbook_preview(path: Path) -> dict[str, Any]:
    preview: dict[str, Any] = {
        "openable": False,
        "sheet_count": 0,
        "sheet_names": [],
    }
    if load_workbook is None:
        return preview
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            sheet_names = list(wb.sheetnames)
            preview["openable"] = True
            preview["sheet_count"] = len(sheet_names)
            preview["sheet_names"] = sheet_names[:8]
        finally:
            wb.close()
    except Exception:
        return preview
    return preview
