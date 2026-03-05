from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from export_excel_grids import EXCEL_PATTERNS
from normalize_rates_to_pg import ALLOWED_SOURCE_COMPANIES, infer_canonical_company_name
from run_update_pipeline import read_manifest

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    from config import DEEPSEEK_API_KEY as CONFIG_DEEPSEEK_API_KEY
    from config import PG_DSN as CONFIG_PG_DSN
except Exception:
    CONFIG_PG_DSN = ""
    CONFIG_DEEPSEEK_API_KEY = ""


LLM_SYSTEM_PROMPT = (
    "You are a strict logistics-file intake classifier for an automated ETL pipeline. "
    "Your task is to decide whether a newly arrived Excel file is a logistics pricing workbook. "
    "If yes, identify the provider, classify the feed, and choose which existing active snapshot file it should replace. "
    "Respond with JSON only."
)
FILENAME_SCAN_KEYWORDS = (*ALLOWED_SOURCE_COMPANIES, "报价")


@dataclass
class IntakeDecision:
    decision: str
    provider: str
    feed_family: str
    replaces_filename: str
    confidence: float
    reason: str


def filename_has_scan_keyword(filename: str) -> bool:
    name = str(filename or "")
    return any(keyword in name for keyword in FILENAME_SCAN_KEYWORDS)


def configure_console_encoding() -> None:
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Watch incoming monthly folders, use LLM filename triage, update update_excel/, and optionally trigger the publish pipeline."
    )
    parser.add_argument(
        "--source-root",
        type=Path,
        default=Path(r"D:\wx_lj\WXWork\1688854674811621\Cache\File"),
        help="Root directory that contains monthly folders like 2026-03, 2026-04",
    )
    parser.add_argument(
        "--update-dir",
        type=Path,
        default=Path.cwd() / "update_excel",
        help="Current active snapshot directory (default: ./update_excel)",
    )
    parser.add_argument(
        "--manifest",
        type=Path,
        default=Path.cwd() / "update_excel" / "manifest.csv",
        help="Manifest file to rewrite after accepted files are imported",
    )
    parser.add_argument(
        "--archive-root",
        type=Path,
        default=Path.cwd() / "archive_excel" / "auto_ingest",
        help="Archive directory for replaced snapshot files",
    )
    parser.add_argument(
        "--state-file",
        type=Path,
        default=Path.cwd() / "out" / "auto_ingest_state.json",
        help="JSON state file used to avoid reprocessing the same files",
    )
    parser.add_argument(
        "--dsn",
        default=os.getenv("PG_DSN", CONFIG_PG_DSN),
        help="PostgreSQL DSN passed through when triggering the pipeline",
    )
    parser.add_argument(
        "--llm-api-key",
        default=os.getenv("DEEPSEEK_API_KEY", CONFIG_DEEPSEEK_API_KEY),
        help="DeepSeek API key for filename triage",
    )
    parser.add_argument(
        "--llm-model",
        default="deepseek-chat",
        help="DeepSeek model used for filename triage",
    )
    parser.add_argument(
        "--llm-confidence-threshold",
        type=float,
        default=0.8,
        help="Minimum confidence required for an automatic accept decision",
    )
    parser.add_argument(
        "--poll-seconds",
        type=float,
        default=20.0,
        help="Polling interval in seconds for watch mode",
    )
    parser.add_argument(
        "--stability-seconds",
        type=float,
        default=2.0,
        help="Seconds to wait before a second size/mtime check to confirm the file has settled",
    )
    parser.add_argument(
        "--debounce-seconds",
        type=float,
        default=120.0,
        help="Quiet period before auto-triggering the publish pipeline after accepted files",
    )
    parser.add_argument(
        "--watch-previous-month",
        dest="watch_previous_month",
        action="store_true",
        help="Also watch the previous month's folder if it exists (default: enabled)",
    )
    parser.add_argument(
        "--no-watch-previous-month",
        dest="watch_previous_month",
        action="store_false",
        help="Only watch the current month's folder",
    )
    parser.add_argument(
        "--run-pipeline",
        dest="run_pipeline",
        action="store_true",
        help="Trigger run_update_pipeline.py after accepted files (default: enabled)",
    )
    parser.add_argument(
        "--no-run-pipeline",
        dest="run_pipeline",
        action="store_false",
        help="Do not trigger the publish pipeline automatically",
    )
    parser.add_argument(
        "--once",
        action="store_true",
        help="Scan once, process stable files, optionally run pipeline once, then exit",
    )
    parser.set_defaults(watch_previous_month=True, run_pipeline=True)
    return parser.parse_args()


def month_folder_name(month_date: date) -> str:
    return month_date.strftime("%Y-%m")


def previous_month(month_date: date) -> date:
    first = month_date.replace(day=1)
    return (first - timedelta(days=1)).replace(day=1)


def determine_watch_dirs(source_root: Path, watch_previous_month: bool) -> list[Path]:
    today = date.today()
    month_names = [month_folder_name(today)]
    if watch_previous_month:
        prev = previous_month(today)
        if month_folder_name(prev) not in month_names:
            month_names.append(month_folder_name(prev))

    watch_dirs: list[Path] = []
    for month_name in month_names:
        path = source_root / month_name
        if path.exists() and path.is_dir():
            watch_dirs.append(path)
    return watch_dirs


def iter_candidate_files(watch_dirs: list[Path]) -> list[Path]:
    files: dict[str, Path] = {}
    for watch_dir in watch_dirs:
        for pattern in EXCEL_PATTERNS:
            for path in watch_dir.glob(pattern):
                if (
                    path.is_file()
                    and not path.name.startswith("~$")
                    and filename_has_scan_keyword(path.name)
                ):
                    files[str(path.resolve())] = path.resolve()
    return sorted(files.values(), key=lambda p: (str(p.parent).lower(), p.name.lower()))


def ensure_state_shape(raw: dict[str, Any] | None) -> dict[str, Any]:
    state = raw or {}
    if not isinstance(state.get("files"), dict):
        state["files"] = {}
    if not isinstance(state.get("processed_hashes"), dict):
        state["processed_hashes"] = {}
    return state


def load_state(path: Path) -> dict[str, Any]:
    if not path.exists():
        return ensure_state_shape({})
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        data = {}
    return ensure_state_shape(data if isinstance(data, dict) else {})


def save_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def iso_now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def file_signature(path: Path) -> tuple[str, int, int]:
    stat = path.stat()
    signature = f"{stat.st_size}:{stat.st_mtime_ns}"
    return signature, int(stat.st_size), int(stat.st_mtime_ns)


def mark_file_seen(path: Path, state: dict[str, Any], now_ts: float) -> dict[str, Any]:
    signature, size, mtime_ns = file_signature(path)
    record = state["files"].setdefault(str(path), {})
    record["signature"] = signature
    record["size"] = size
    record["mtime_ns"] = mtime_ns
    record["stable_since"] = now_ts
    record["last_seen_at"] = iso_now()
    return record


def confirm_file_settled(path: Path, record: dict[str, Any], recheck_seconds: float) -> bool:
    expected_signature = str(record.get("signature") or "")
    delay = max(0.0, float(recheck_seconds))
    if delay > 0:
        time.sleep(delay)

    try:
        signature, size, mtime_ns = file_signature(path)
    except Exception as exc:
        record["last_error"] = f"restat_failed:{exc}"
        return False

    record["last_seen_at"] = iso_now()
    if signature != expected_signature:
        record["signature"] = signature
        record["size"] = size
        record["mtime_ns"] = mtime_ns
        record["stable_since"] = time.time()
        record["last_error"] = "file_changed_during_recheck"
        return False

    if not file_readable(path):
        record["last_error"] = "file_not_readable"
        return False

    record["last_error"] = ""
    return True


def file_readable(path: Path) -> bool:
    try:
        with path.open("rb"):
            return True
    except Exception:
        return False


def compute_file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def load_workbook_preview(path: Path) -> dict[str, Any]:
    preview: dict[str, Any] = {
        "openable": False,
        "sheet_count": 0,
        "sheet_names": [],
    }
    if load_workbook is None:
        return preview
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            sheet_names = list(wb.sheetnames)
            preview["openable"] = True
            preview["sheet_count"] = len(sheet_names)
            preview["sheet_names"] = sheet_names[:8]
        finally:
            wb.close()
    except Exception:
        return preview
    return preview


def parse_json_object(raw: str) -> dict[str, Any] | None:
    text = raw.strip()
    if not text:
        return None
    try:
        data = json.loads(text)
        return data if isinstance(data, dict) else None
    except Exception:
        pass
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        try:
            data = json.loads(text[start : end + 1])
            return data if isinstance(data, dict) else None
        except Exception:
            return None
    return None


def build_llm_client(api_key: str) -> Any:
    if OpenAI is None:
        raise SystemExit("openai package is missing. Install it before using the auto-ingest watcher.")
    if not api_key:
        raise SystemExit("Missing LLM API key. Set DEEPSEEK_API_KEY or pass --llm-api-key.")
    return OpenAI(api_key=api_key, base_url="https://api.deepseek.com")


def read_active_manifest_rows(manifest_path: Path) -> list[dict[str, Any]]:
    if manifest_path.exists():
        return [row for row in read_manifest(manifest_path) if row.get("enabled")]
    return []


def current_snapshot_filenames(update_dir: Path, manifest_path: Path) -> list[str]:
    rows = read_active_manifest_rows(manifest_path)
    if rows:
        return [str(row["filename"]) for row in rows]

    files: list[str] = []
    for pattern in EXCEL_PATTERNS:
        files.extend([p.name for p in update_dir.glob(pattern) if p.is_file() and not p.name.startswith("~$")])
    return sorted(dict.fromkeys(files))


def classify_with_llm(
    client: Any,
    model: str,
    path: Path,
    workbook_preview: dict[str, Any],
    current_snapshot: list[str],
) -> IntakeDecision:
    payload = {
        "task": "classify_incoming_logistics_pricing_workbook",
        "filename": path.name,
        "file_stem": path.stem,
        "folder_path": str(path.parent),
        "allowed_providers": list(ALLOWED_SOURCE_COMPANIES),
        "current_snapshot_filenames": current_snapshot,
        "workbook_preview": workbook_preview,
        "rules": [
            "Ignore irrelevant files that are not logistics pricing workbooks.",
            "If accepted, provider must be one of the allowed providers.",
            "replaces_filename must be exactly one existing current snapshot filename or empty string.",
            "Prefer replace_existing when the file is an updated version of an existing feed.",
            "Use add_new_feed when it is a new logistics feed.",
        ],
        "output_schema": {
            "decision": "accept|ignore|review",
            "provider": "one of allowed providers or unknown",
            "feed_family": "short identifier string",
            "replaces_filename": "exact existing filename or empty string",
            "confidence": "number_0_to_1",
            "reason": "short string",
        },
    }

    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": LLM_SYSTEM_PROMPT},
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
        ],
        temperature=0.0,
        stream=False,
    )
    content = response.choices[0].message.content or ""
    parsed = parse_json_object(content)
    if parsed is None:
        return IntakeDecision(
            decision="review",
            provider="unknown",
            feed_family="",
            replaces_filename="",
            confidence=0.0,
            reason="llm_parse_failed",
        )

    decision = str(parsed.get("decision") or "").strip().lower()
    if decision not in {"accept", "ignore", "review"}:
        decision = "review"

    provider = str(parsed.get("provider") or "").strip()
    feed_family = str(parsed.get("feed_family") or "").strip()
    replaces_filename = str(parsed.get("replaces_filename") or "").strip()
    if replaces_filename not in current_snapshot:
        replaces_filename = ""

    confidence_raw = parsed.get("confidence", 0)
    try:
        confidence = float(confidence_raw)
    except Exception:
        confidence = 0.0
    confidence = max(0.0, min(1.0, confidence))

    reason = str(parsed.get("reason") or "").strip()

    return IntakeDecision(
        decision=decision,
        provider=provider,
        feed_family=feed_family,
        replaces_filename=replaces_filename,
        confidence=confidence,
        reason=reason,
    )


def canonicalize_decision(
    decision: IntakeDecision,
    filename: str,
    confidence_threshold: float,
) -> IntakeDecision:
    deterministic_provider = infer_canonical_company_name(filename)
    provider = decision.provider if decision.provider in ALLOWED_SOURCE_COMPANIES else "unknown"
    final_decision = decision.decision
    reason_parts: list[str] = [decision.reason] if decision.reason else []

    if provider == "unknown":
        final_decision = "review"
        reason_parts.append("provider_not_allowed")

    if deterministic_provider and provider != "unknown" and deterministic_provider != provider:
        final_decision = "review"
        reason_parts.append(f"provider_mismatch:{deterministic_provider}")

    if final_decision == "accept" and decision.confidence < confidence_threshold:
        final_decision = "review"
        reason_parts.append("confidence_below_threshold")

    if final_decision == "accept" and not decision.feed_family:
        final_decision = "review"
        reason_parts.append("feed_family_missing")

    return IntakeDecision(
        decision=final_decision,
        provider=provider,
        feed_family=decision.feed_family,
        replaces_filename=decision.replaces_filename,
        confidence=decision.confidence,
        reason=" | ".join([part for part in reason_parts if part]),
    )


def archive_file(path: Path, archive_root: Path) -> Path:
    archive_root.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = archive_root / f"{stamp}_{path.name}"
    counter = 1
    while candidate.exists():
        candidate = archive_root / f"{stamp}_{counter}_{path.name}"
        counter += 1
    path.replace(candidate)
    return candidate


def load_manifest_map(manifest_path: Path) -> dict[str, dict[str, Any]]:
    rows = read_active_manifest_rows(manifest_path)
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        result[str(row["filename"])] = {
            "filename": str(row["filename"]),
            "enabled": bool(row["enabled"]),
            "expect_channels": bool(row["expect_channels"]),
            "notes": str(row.get("notes") or ""),
        }
    return result


def write_manifest_map(manifest_path: Path, rows_by_name: dict[str, dict[str, Any]]) -> None:
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["filename", "enabled", "expect_channels", "notes"]
    with manifest_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for filename in sorted(rows_by_name, key=lambda x: x.lower()):
            row = rows_by_name[filename]
            writer.writerow(
                {
                    "filename": row["filename"],
                    "enabled": "1" if row.get("enabled", True) else "0",
                    "expect_channels": "1" if row.get("expect_channels", True) else "0",
                    "notes": row.get("notes", ""),
                }
            )


def file_hash_or_empty(path: Path) -> str:
    try:
        return compute_file_hash(path)
    except Exception:
        return ""


def apply_snapshot_update(
    incoming_path: Path,
    decision: IntakeDecision,
    update_dir: Path,
    manifest_path: Path,
    archive_root: Path,
) -> dict[str, Any]:
    update_dir.mkdir(parents=True, exist_ok=True)
    archive_root.mkdir(parents=True, exist_ok=True)

    manifest_map = load_manifest_map(manifest_path)
    replaced_filename = decision.replaces_filename

    if replaced_filename:
        existing = update_dir / replaced_filename
        if existing.exists():
            archive_file(existing, archive_root)
        manifest_map.pop(replaced_filename, None)

    destination = update_dir / incoming_path.name
    if destination.exists():
        existing_hash = file_hash_or_empty(destination)
        incoming_hash = file_hash_or_empty(incoming_path)
        if existing_hash and incoming_hash and existing_hash == incoming_hash:
            note = f"auto-ingest noop {date.today().isoformat()} feed={decision.feed_family}"
            manifest_map[incoming_path.name] = {
                "filename": incoming_path.name,
                "enabled": True,
                "expect_channels": True,
                "notes": note,
            }
            write_manifest_map(manifest_path, manifest_map)
            return {
                "status": "noop",
                "destination": str(destination),
                "replaced_filename": replaced_filename,
            }
        archive_file(destination, archive_root)

    import shutil

    shutil.copy2(incoming_path, destination)
    manifest_note = (
        f"auto-ingested {date.today().isoformat()} "
        f"provider={decision.provider} feed={decision.feed_family}"
    )
    if replaced_filename:
        manifest_note += f" replace={replaced_filename}"
    else:
        manifest_note += " new_feed"
    manifest_map[incoming_path.name] = {
        "filename": incoming_path.name,
        "enabled": True,
        "expect_channels": True,
        "notes": manifest_note,
    }
    write_manifest_map(manifest_path, manifest_map)
    return {
        "status": "copied",
        "destination": str(destination),
        "replaced_filename": replaced_filename,
    }


def run_publish_pipeline(args: argparse.Namespace, script_dir: Path) -> None:
    if not args.dsn:
        raise SystemExit("Missing DSN. Provide --dsn or set PG_DSN if you want to auto-run the pipeline.")

    cmd = [
        sys.executable,
        "run_update_pipeline.py",
        "--manifest",
        str(args.manifest.resolve()),
        "--manifest-required",
        "--dsn",
        args.dsn,
        "--llm-model",
        args.llm_model,
    ]
    if args.llm_api_key:
        cmd.extend(["--llm-api-key", args.llm_api_key])

    env = os.environ.copy()
    env.setdefault("PYTHONUTF8", "1")
    env.setdefault("PYTHONIOENCODING", "utf-8")
    proc = subprocess.run(
        cmd,
        cwd=str(script_dir),
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    print(proc.stdout, end="")
    if proc.returncode != 0:
        raise RuntimeError(f"run_update_pipeline.py failed with exit code {proc.returncode}")


def process_file(
    path: Path,
    record: dict[str, Any],
    state: dict[str, Any],
    args: argparse.Namespace,
    client: Any,
) -> tuple[bool, str]:
    file_hash = compute_file_hash(path)
    record["file_hash"] = file_hash

    existing_hash = state["processed_hashes"].get(file_hash)
    if existing_hash and str(existing_hash.get("status") or "") in {"accepted", "ignored", "review"}:
        record["status"] = "duplicate"
        record["last_processed_signature"] = record.get("signature", "")
        record["last_processed_at"] = iso_now()
        record["decision"] = {
            "decision": "duplicate",
            "provider": str(existing_hash.get("provider") or ""),
            "reason": "same_hash_already_processed",
        }
        return False, "duplicate"

    workbook_preview = load_workbook_preview(path)
    current_snapshot = current_snapshot_filenames(args.update_dir, args.manifest)
    raw_decision = classify_with_llm(client, args.llm_model, path, workbook_preview, current_snapshot)
    decision = canonicalize_decision(raw_decision, path.name, float(args.llm_confidence_threshold))

    record["decision"] = {
        "decision": decision.decision,
        "provider": decision.provider,
        "feed_family": decision.feed_family,
        "replaces_filename": decision.replaces_filename,
        "confidence": decision.confidence,
        "reason": decision.reason,
    }
    record["last_processed_signature"] = record.get("signature", "")
    record["last_processed_at"] = iso_now()

    if decision.decision != "accept":
        record["status"] = decision.decision
        state["processed_hashes"][file_hash] = {
            "status": decision.decision,
            "provider": decision.provider,
            "filename": path.name,
            "processed_at": record["last_processed_at"],
        }
        return False, decision.decision

    update_result = apply_snapshot_update(
        incoming_path=path,
        decision=decision,
        update_dir=args.update_dir,
        manifest_path=args.manifest,
        archive_root=args.archive_root,
    )
    record["status"] = "accepted"
    record["update_result"] = update_result
    state["processed_hashes"][file_hash] = {
        "status": "accepted",
        "provider": decision.provider,
        "filename": path.name,
        "processed_at": record["last_processed_at"],
        "replaces_filename": decision.replaces_filename,
    }
    return True, update_result.get("status", "accepted")


def scan_once(
    args: argparse.Namespace,
    client: Any,
) -> tuple[int, int]:
    state = load_state(args.state_file)
    now_ts = time.time()
    watch_dirs = determine_watch_dirs(args.source_root, bool(args.watch_previous_month))
    if not watch_dirs:
        print(f"No watch folders found under {args.source_root}")
        save_state(args.state_file, state)
        return 0, 0

    print("Watching folders:")
    for path in watch_dirs:
        print(f"  - {path}")

    accepted_count = 0
    scanned_count = 0
    recheck_seconds = float(args.stability_seconds)
    for path in iter_candidate_files(watch_dirs):
        scanned_count += 1
        record = mark_file_seen(path, state, now_ts)

        if str(record.get("last_processed_signature") or "") == str(record.get("signature") or ""):
            continue

        if not confirm_file_settled(path, record, recheck_seconds):
            continue

        try:
            accepted, result = process_file(path, record, state, args, client)
            print(f"[{record.get('status','?')}] {path.name} -> {result}")
            if accepted:
                accepted_count += 1
        except Exception as exc:
            record["status"] = "failed"
            record["last_error"] = str(exc)
            record["last_processed_signature"] = record.get("signature", "")
            record["last_processed_at"] = iso_now()
            print(f"[failed] {path.name} -> {exc}")

    save_state(args.state_file, state)
    return scanned_count, accepted_count


def main() -> None:
    configure_console_encoding()
    args = parse_args()
    args.source_root = args.source_root.resolve()
    args.update_dir = args.update_dir.resolve()
    args.manifest = args.manifest.resolve()
    args.archive_root = args.archive_root.resolve()
    args.state_file = args.state_file.resolve()

    threshold = max(0.0, min(1.0, float(args.llm_confidence_threshold)))
    args.llm_confidence_threshold = threshold
    client = build_llm_client(args.llm_api_key)
    script_dir = Path(__file__).resolve().parent

    pending_pipeline = False
    next_pipeline_at: float | None = None

    while True:
        scanned_count, accepted_count = scan_once(args, client)
        print(f"Scan complete: scanned={scanned_count}, accepted={accepted_count}")

        if accepted_count > 0 and args.run_pipeline:
            if args.once:
                run_publish_pipeline(args, script_dir)
            else:
                pending_pipeline = True
                next_pipeline_at = time.time() + float(args.debounce_seconds)
                print(
                    "Pipeline run scheduled "
                    f"in {float(args.debounce_seconds):.0f}s after quiet period."
                )

        if args.once:
            break

        now_ts = time.time()
        if pending_pipeline and next_pipeline_at is not None and now_ts >= next_pipeline_at:
            try:
                run_publish_pipeline(args, script_dir)
                pending_pipeline = False
                next_pipeline_at = None
            except Exception as exc:
                print(f"Pipeline trigger failed: {exc}")
                next_pipeline_at = time.time() + float(args.debounce_seconds)

        time.sleep(max(1.0, float(args.poll_seconds)))


if __name__ == "__main__":
    main()
